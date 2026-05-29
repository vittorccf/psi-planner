# -*- coding: utf-8 -*-
"""
Build da planilha Psi Planner (estrutura .xlsx) a partir do visual_spec.json.
Estilo vem do spec; dados/fórmulas/named ranges/validações/CF/gráficos são
adicionados aqui. Fórmulas nativas mantêm tudo vivo sem macros.
"""
import json
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.title import Title
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.text import (Paragraph, ParagraphProperties,
                                   CharacterProperties, Font as DrawFont, RegularTextRun)
from openpyxl.drawing.line import LineProperties
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

BUILD = r"D:\Vittor\Projetos\Saluti\PsiPlanner\build"
OUT_XLSX = r"D:\Vittor\Projetos\Saluti\PsiPlanner\build\PsiPlanner_base.xlsx"
PRODUTO = "Psi Planner"  # nome do produto (planner financeiro-clínico)

with open(BUILD + r"\visual_spec.json", encoding="utf-8") as f:
    SPEC = json.load(f)

PAL = SPEC["meta"]["palette"]
FONTS = SPEC["fonts"]

def hexc(key_or_hex):
    """Aceita chave da paleta ou HEX direto -> retorna HEX 6 chars."""
    if key_or_hex is None:
        return None
    return PAL.get(key_or_hex, key_or_hex)

def argb(h):
    return "FF" + h if h and len(h) == 6 else h

def mkfont(font_key, color_hex):
    fd = FONTS[font_key]
    return Font(name=fd["applied"], size=fd["size"], bold=fd.get("bold", False),
                italic=fd.get("italic", False), color=argb(color_hex or "000000"))

def side(style, color):
    return Side(style=style, color=argb(color))

ALIGN = lambda h, v, wrap=False: Alignment(horizontal=h, vertical=v, wrap_text=wrap)

wb = Workbook()
wb.remove(wb.active)
sheets = {}
for name in SPEC["meta"]["sheet_order"]:
    ws = wb.create_sheet(title=name)
    sheets[name] = ws

# ---------------------------------------------------------------------------
# PASSO 1 — aplicar estilo do spec (dimensões, células, merges, bordas)
# ---------------------------------------------------------------------------
def apply_borders(ws, b):
    # Redesign: respiro > moldura. Sem caixas (outline); só hairline clara
    # horizontal. "all" (grade) vira só linhas horizontais.
    edge = b["edge"]
    if edge == "outline":
        return
    s = side("thin", PAL["line"])
    rows = list(ws[b["range"]])
    nrows = len(rows)
    if edge == "all":
        edge = "hlines"
    for ri, row in enumerate(rows):
        for cell in row:
            top = bottom = None
            if edge == "hlines":
                bottom = s
            elif edge == "bottom":
                if ri == nrows - 1: bottom = s
            elif edge == "top":
                if ri == 0: top = s
            ex = cell.border
            cell.border = Border(top=top or ex.top, left=ex.left,
                                 right=ex.right, bottom=bottom or ex.bottom)

for name, sdef in SPEC["sheets"].items():
    ws = sheets[name]
    for col, w in sdef.get("column_widths", {}).items():
        ws.column_dimensions[col].width = w
    for row, h in sdef.get("row_heights", {}).items():
        ws.row_dimensions[int(row)].height = h
    for c in sdef.get("cells", []):
        cell = ws[c["ref"]]
        cell.value = c["value"]
        cell.font = mkfont(c["font"], c.get("font_color"))
        fill = hexc(c.get("fill"))
        if fill:
            cell.fill = PatternFill(fill_type="solid", fgColor=argb(fill))
        cell.alignment = ALIGN(c.get("align", "left"), c.get("valign", "center"),
                               wrap=c.get("wrap", False))
    for m in sdef.get("merges", []):
        try:
            ws.merge_cells(m)
        except Exception:
            pass
    for b in sdef.get("borders", []):
        apply_borders(ws, b)

# ---------------------------------------------------------------------------
# PASSO 2 — named ranges (mapa autoritativo, escopo workbook)
# ---------------------------------------------------------------------------
def q(sheet):  # cita nome de aba para fórmula/refersto
    return "'" + sheet + "'"

NAMED = {
    # Menu KPIs -> células de VALOR (linha 7)
    "KPI_PacientesAtivos": ("Menu Inicial", "$B$7"),
    "KPI_ReceitaMes": ("Menu Inicial", "$C$7"),
    "KPI_ValorPendente": ("Menu Inicial", "$D$7"),
    "KPI_ProxConsulta": ("Menu Inicial", "$F$7"),
    "KPI_FaltasMes": ("Menu Inicial", "$G$7"),
    "KPI_AniversariantesMes": ("Menu Inicial", "$H$7"),
    # Avisos (4 blocos)
    "AVISO_SemSessao": ("Menu Inicial", "$J$6"),
    "AVISO_MEI": ("Menu Inicial", "$J$9"),
    "AVISO_Pendentes": ("Menu Inicial", "$J$11"),
    "AVISO_Aniversariantes": ("Menu Inicial", "$J$13"),
    # Financeiro lançamentos
    "FIN_Data": ("Financeiro", "$B$7:$B$56"),
    "FIN_Paciente": ("Financeiro", "$C$7:$C$56"),
    "FIN_Tipo": ("Financeiro", "$D$7:$D$56"),
    "FIN_Forma": ("Financeiro", "$E$7:$E$56"),
    "FIN_Valor": ("Financeiro", "$F$7:$F$56"),
    "FIN_Status": ("Financeiro", "$G$7:$G$56"),
    "FIN_NF": ("Financeiro", "$H$7:$H$56"),
    "FIN_NumNF": ("Financeiro", "$I$7:$I$56"),
    "FIN_Obs": ("Financeiro", "$J$7:$J$56"),
    # Financeiro despesas
    "DESP_Data": ("Financeiro", "$B$61:$B$74"),
    "DESP_Categoria": ("Financeiro", "$C$61:$C$74"),
    "DESP_Desc": ("Financeiro", "$D$61:$D$74"),
    "DESP_Valor": ("Financeiro", "$E$61:$E$74"),
    "DESP_Forma": ("Financeiro", "$F$61:$F$74"),
    # Financeiro KPIs
    "FIN_TotalRecebido": ("Financeiro", "$M$6"),
    "FIN_TotalPendente": ("Financeiro", "$M$7"),
    "FIN_LucroLiquido": ("Financeiro", "$M$8"),
    "FIN_Inadimplencia": ("Financeiro", "$M$9"),
    "MEI_Indicador": ("Financeiro", "$M$10"),
    "MEI_Faturamento": ("Financeiro", "$M$11"),
    # Pacientes
    "PAC_Nome": ("Pacientes", "$B$7:$B$56"),
    "PAC_Nasc": ("Pacientes", "$E$7:$E$56"),
    "PAC_Idade": ("Pacientes", "$F$7:$F$56"),
    "PAC_Convenio": ("Pacientes", "$J$7:$J$56"),
    "PAC_ValorSessao": ("Pacientes", "$N$7:$N$56"),
    "PAC_Status": ("Pacientes", "$O$7:$O$56"),
    "PAC_SessContratadas": ("Pacientes", "$P$7:$P$56"),
    "PAC_SessRealizadas": ("Pacientes", "$Q$7:$Q$56"),
    "PAC_Saldo": ("Pacientes", "$R$7:$R$56"),
    "PAC_UltimaSessao": ("Pacientes", "$S$7:$S$56"),
    # Calendário
    "CAL_Grid": ("Calendário", "$C$6:$H$21"),
    "CAL_ConsultasDia": ("Calendário", "$K$5"),
    "CAL_TaxaCancelamento": ("Calendário", "$K$6"),
    "CAL_TaxaFaltas": ("Calendário", "$K$7"),
    "CAL_LogData": ("Calendário", "$A$100:$A$160"),
    "CAL_LogStatus": ("Calendário", "$E$100:$E$160"),
    # Aniversariantes
    "ANIV_Nome": ("Aniversariantes", "$B$7:$B$56"),
    "ANIV_Nasc": ("Aniversariantes", "$C$7:$C$56"),
    "ANIV_Idade": ("Aniversariantes", "$D$7:$D$56"),
    "ANIV_Telefone": ("Aniversariantes", "$F$7:$F$56"),
    "ANIV_Status": ("Aniversariantes", "$G$7:$G$56"),
    "ANIV_MsgEnviada": ("Aniversariantes", "$H$7:$H$56"),
    "ANIV_TotalMes": ("Aniversariantes", "$K$5"),
    # Relatórios
    "REL_ReceitaMensal": ("Relatórios", "$C$6"),
    "REL_ReceitaAnual": ("Relatórios", "$C$7"),
    "REL_Atendimentos": ("Relatórios", "$C$8"),
    "REL_Ativos": ("Relatórios", "$F$6"),
    "REL_TaxaCancelamento": ("Relatórios", "$F$7"),
    "REL_TaxaInadimplencia": ("Relatórios", "$F$8"),
    "REL_ReceitaConvenio": ("Relatórios", "$I$6"),
    "REL_ComparativoMeses": ("Relatórios", "$I$7"),
    # Configurações
    "CFG_NomeClinica": ("Configurações", "$C$6"),
    "CFG_Profissional": ("Configurações", "$C$7"),
    "CFG_Registro": ("Configurações", "$C$8"),
    "CFG_Telefone": ("Configurações", "$C$9"),
    "CFG_Email": ("Configurações", "$C$10"),
    "CFG_LimiteMEI": ("Configurações", "$C$13"),
    "CFG_ValorSessao": ("Configurações", "$C$14"),
    "CFG_ExibirConvenio": ("Configurações", "$C$18"),
    "CFG_Tema": ("Configurações", "$C$21"),
}
for nm, (sh, ref) in NAMED.items():
    wb.defined_names[nm] = DefinedName(nm, attr_text=q(sh) + "!" + ref)

# ---------------------------------------------------------------------------
# PASSO 3 — dados seed (clínica-exemplo)
# ---------------------------------------------------------------------------
D = dt.date  # atalho

# Pacientes: (Nome,CPF,RG,Nasc,Tel,Email,Endereço,Convênio,Profissão,Início,Freq,Valor,Status,Contr,Realiz,UltSessao)
PACIENTES = [
    ("Ana Beatriz Souza","111.111.111-11","11.111.111-1",D(1990,3,10),"(62) 98111-2222","ana.souza@email.com","R. das Flores, 120 - Goiânia/GO","Particular","Arquiteta",D(2025,9,15),"Semanal",180,"Ativo",12,9,D(2026,5,4)),
    ("Bruno Carvalho","222.222.222-22","22.222.222-2",D(1988,5,15),"(62) 98222-3333","bruno.c@email.com","Av. T-63, 800 - Goiânia/GO","Unimed","Engenheiro",D(2025,11,2),"Semanal",180,"Ativo",10,7,D(2026,5,6)),
    ("Camila Diniz","333.333.333-33","33.333.333-3",D(1992,5,28),"(62) 98333-4444","camila.d@email.com","R. 7, 45 - Anápolis/GO","Particular","Professora",D(2026,1,20),"Quinzenal",150,"Ativo",8,3,D(2026,5,11)),
    ("Diego Esteves","444.444.444-44","44.444.444-4",D(1995,7,8),"(62) 98444-5555","diego.e@email.com","R. 9, 300 - Goiânia/GO","Particular","Designer",D(2026,2,12),"Semanal",200,"Pausado",6,4,D(2026,4,15)),
    ("Eduarda Faria","555.555.555-55","55.555.555-5",D(1991,5,30),"(62) 98555-6666","edu.faria@email.com","Av. 85, 1500 - Goiânia/GO","Bradesco Saúde","Médica",D(2025,8,25),"Semanal",180,"Ativo",20,16,D(2026,5,18)),
    ("Fernando Gomes","666.666.666-66","66.666.666-6",D(1989,6,12),"(62) 98666-7777","fer.gomes@email.com","R. C-140, 22 - Goiânia/GO","Particular","Advogado",D(2025,12,1),"Quinzenal",180,"Ativo",10,5,D(2026,5,22)),
    ("Gabriela Hoffmann","777.777.777-77","77.777.777-7",D(1994,9,25),"(62) 98777-8888","gabi.h@email.com","R. 1004, 60 - Goiânia/GO","SulAmérica","Publicitária",D(2026,3,3),"Semanal",180,"Ativo",8,6,D(2026,5,25)),
    ("Kris Fellipe","888.888.888-88","88.888.888-8",D(1986,6,3),"(62) 98888-9999","kris.fellipe@email.com","UBS Turvânia - Turvânia/GO","Particular","Dentista (UBS)",D(2026,1,10),"Mensal",150,"Ativo",6,2,D(2026,5,20)),
]

# Financeiro lançamentos: (Data,Paciente,Tipo,Forma,Valor,Status,NF,NumNF,Obs)
FIN = [
    (D(2026,1,20),"Ana Beatriz Souza","Sessão presencial","Pix",180,"Pago","Emitida","2026-091","Recibo Receita Saúde"),
    (D(2026,2,18),"Bruno Carvalho","Sessão online","Cartão",180,"Pago","Emitida","2026-098",""),
    (D(2026,3,17),"Camila Diniz","Sessão presencial","Dinheiro",150,"Pago","Não emitida","","Pagamento em espécie"),
    (D(2026,4,14),"Eduarda Faria","Sessão online","Pix",180,"Pago","Emitida","2026-110",""),
    (D(2026,4,15),"Diego Esteves","Sessão presencial","Pix",180,"Cancelado","Não emitida","","Cancelou 24h antes"),
    (D(2026,5,4),"Ana Beatriz Souza","Sessão presencial","Pix",180,"Pago","Emitida","2026-121","Recibo Receita Saúde"),
    (D(2026,5,6),"Bruno Carvalho","Sessão online","Cartão",180,"Pago","Emitida","2026-122",""),
    (D(2026,5,11),"Camila Diniz","Sessão presencial","Dinheiro",150,"Pago","Não emitida","","Pagamento em espécie"),
    (D(2026,5,13),"Diego Esteves","Sessão presencial","Pix",200,"Pendente","Não emitida","","Aguardando confirmação"),
    (D(2026,5,18),"Eduarda Faria","Sessão online","Pix",180,"Pago","Emitida","2026-123",""),
    (D(2026,5,20),"Kris Fellipe","Sessão presencial","Convênio",150,"Pendente","Não emitida","","Faturamento convênio"),
    (D(2026,5,22),"Fernando Gomes","Sessão presencial","Cartão",180,"Pago","Emitida","2026-124",""),
    (D(2026,5,25),"Gabriela Hoffmann","Sessão online","Pix",180,"Lembrete","Não emitida","","Lembrete WhatsApp enviado"),
]

# Despesas: (Data,Categoria,Descrição,Valor,Forma)
DESP = [
    (D(2026,5,2),"Plataforma online","Assinatura Psi Planner Pro",99,"Cartão"),
    (D(2026,5,3),"Internet","Banda larga + WhatsApp Business",110,"Boleto"),
    (D(2026,5,5),"Aluguel","Sala em coworking clínico",600,"Pix"),
    (D(2026,5,8),"Supervisão","Supervisão clínica mensal",300,"Pix"),
    (D(2026,5,10),"Marketing","Google Ads + Instagram",150,"Cartão"),
    (D(2026,5,12),"Materiais","Testes psicométricos e prontuários",120,"Pix"),
    (D(2026,5,15),"Outros","Manutenção de equipamentos",80,"Dinheiro"),
]

# Calendário: grid (dia_col, hora_row, paciente, tipo, status) + log
# colunas: C=Seg D=Ter E=Qua F=Qui G=Sex H=Sáb ; linhas 6..21 = 07h..22h
CAL = [
    ("C",13,"Ana Beatriz Souza","Presencial","Realizado",  D(2026,5,25),"14:00"),
    ("D", 9,"Bruno Carvalho","Online","Confirmado",        D(2026,5,26),"10:00"),
    ("E",14,"Camila Diniz","Presencial","Confirmado",       D(2026,5,27),"15:00"),
    ("F", 8,"Diego Esteves","Presencial","Faltou",          D(2026,5,28),"09:00"),
    ("G",15,"Eduarda Faria","Online","Confirmado",          D(2026,5,29),"16:00"),
    ("H", 7,"Kris Fellipe","Presencial","Cancelado",        D(2026,5,30),"08:00"),
    ("C",10,"Fernando Gomes","Presencial","Confirmado",     D(2026,5,25),"11:00"),
    ("E",18,"Gabriela Hoffmann","Online","Realizado",       D(2026,5,27),"19:00"),
]
ESPERA = [("Marina Lopes","(62) 99100-2200"), ("Otávio Reis","(62) 99100-3300")]

NF_BORDER = side("thin", PAL["line"])  # régua de linha clara (redesign)
def style_data_cell(cell, font_key="conteudo", color="000000", align="left", numfmt=None):
    cell.font = mkfont(font_key, color)
    cell.alignment = ALIGN(align, "center")
    cell.border = Border(bottom=NF_BORDER)
    if numfmt:
        cell.number_format = numfmt

# --- preencher Pacientes ---
ws = sheets["Pacientes"]
r0 = 7
for i, p in enumerate(PACIENTES):
    r = r0 + i
    (nome,cpf,rg,nasc,tel,email,end,conv,prof,inicio,freq,valor,status,contr,realiz,ult) = p
    vals = {
        "B": nome, "C": cpf, "D": rg, "E": nasc, "F": None, "G": tel, "H": email,
        "I": end, "J": conv, "K": prof, "L": inicio, "M": freq, "N": valor,
        "O": status, "P": contr, "Q": realiz, "R": None, "S": ult, "T": None,
    }
    for col, v in vals.items():
        c = ws[f"{col}{r}"]
        if v is not None:
            c.value = v
        algn = "left" if col in ("B","H","I","K") else "center"
        nf = None
        if col in ("E","L","S"):
            nf = "dd/mm/yyyy"
        if col in ("N","T"):
            nf = 'R$ #,##0.00'; algn = "right"
        style_data_cell(c, color="000000", align=algn, numfmt=nf)
    ws[f"F{r}"].value = f"=IF(E{r}=\"\",\"\",DATEDIF(E{r},TODAY(),\"Y\"))"
    ws[f"R{r}"].value = f"=IF(P{r}=\"\",\"\",P{r}-Q{r})"
    ws[f"T{r}"].value = f"=IF(N{r}=\"\",\"\",Q{r}*N{r})"

# --- preencher Aniversariantes (mesmos pacientes) ---
ws = sheets["Aniversariantes"]
for i, p in enumerate(PACIENTES):
    r = r0 + i
    nome, nasc, tel, status = p[0], p[3], p[4], p[12]
    ws[f"B{r}"].value = nome
    ws[f"C{r}"].value = nasc
    ws[f"E{r}"].value = f'=IF(C{r}="","",TEXT(C{r},"dd/mm"))'
    ws[f"F{r}"].value = tel
    ws[f"G{r}"].value = status
    ws[f"H{r}"].value = "Não"
    ws[f"D{r}"].value = f'=IF(C{r}="","",DATEDIF(C{r},TODAY(),"Y"))'
    for col, algn in [("B","left"),("C","center"),("D","center"),("E","center"),("F","center"),("G","center"),("H","center")]:
        nf = "dd/mm/yyyy" if col == "C" else None
        style_data_cell(ws[f"{col}{r}"], color="000000", align=algn, numfmt=nf)
# duas mensagens marcadas como enviadas (demo)
ws["H7"].value = "Sim"
ws["H9"].value = "Sim"

# --- preencher Financeiro lançamentos ---
ws = sheets["Financeiro"]
for i, t in enumerate(FIN):
    r = 7 + i
    data,pac,tipo,forma,valor,status,nf,numnf,obs = t
    cells = {"B":data,"C":pac,"D":tipo,"E":forma,"F":valor,"G":status,"H":nf,"I":numnf,"J":obs}
    for col, v in cells.items():
        c = ws[f"{col}{r}"]
        c.value = v
        algn = "center"; numfmt = None
        if col in ("C","J"): algn = "left"
        if col == "B": numfmt = "dd/mm/yyyy"
        if col == "F": numfmt = 'R$ #,##0.00'; algn = "right"
        style_data_cell(c, color="000000", align=algn, numfmt=numfmt)

# --- preencher Despesas ---
for i, d in enumerate(DESP):
    r = 61 + i
    data,cat,desc,valor,forma = d
    cells = {"B":data,"C":cat,"D":desc,"E":valor,"F":forma}
    for col, v in cells.items():
        c = ws[f"{col}{r}"]
        c.value = v
        algn = "center"; numfmt = None
        if col in ("C","D"): algn = "left"
        if col == "B": numfmt = "dd/mm/yyyy"
        if col == "E": numfmt = 'R$ #,##0.00'; algn = "right"
        style_data_cell(c, color="000000", align=algn, numfmt=numfmt)

# --- preencher Calendário grid + log ---
ws = sheets["Calendário"]
grid_fill = PatternFill(fill_type="solid", fgColor=argb(PAL["branco"]))
for (col, row, pac, tipo, status, data, hora) in CAL:
    c = ws[f"{col}{row}"]
    c.value = f"{pac}\n{tipo} · {status}"
    c.font = Font(name="Inter", size=8, color=argb("000000"))  # menor: cabe sem cortar
    c.alignment = ALIGN("center", "center", wrap=True)
# Registro de agenda (log) - cabeçalho linha 99, dados 100+
ws["B97"].value = "Registro de agenda (base de dados)"
ws["B97"].font = mkfont("conteudo_pequeno", "A6A6A6")
hdr = {"A":"Data","B":"Hora","C":"Paciente","D":"Tipo","E":"Status"}
for col, v in hdr.items():
    cc = ws[f"{col}99"]
    cc.value = v
    cc.font = mkfont("conteudo_pequeno", "3B3B3B")
for i, (col, row, pac, tipo, status, data, hora) in enumerate(CAL):
    lr = 100 + i
    ws[f"A{lr}"].value = data; ws[f"A{lr}"].number_format = "dd/mm/yyyy"
    ws[f"B{lr}"].value = hora
    ws[f"C{lr}"].value = pac
    ws[f"D{lr}"].value = tipo
    ws[f"E{lr}"].value = status
    for col in ("A","B","C","D","E"):
        ws[f"{col}{lr}"].font = mkfont("conteudo_pequeno", "3B3B3B")
# coluna auxiliar (AB) p/ "próxima consulta": data se futura, senão vazio
for r in range(100, 161):
    ws[f"AB{r}"] = f'=IF(AND(A{r}<>"",A{r}>=TODAY()),A{r},"")'
# lista de espera
for i, (nome, tel) in enumerate(ESPERA):
    r = 12 + i
    ws[f"J{r}"].value = nome; ws[f"J{r}"].font = mkfont("conteudo", "000000")
    ws[f"K{r}"].value = tel; ws[f"K{r}"].font = mkfont("conteudo", "000000")
    ws[f"J{r}"].alignment = ALIGN("left","center"); ws[f"K{r}"].alignment = ALIGN("center","center")

# ---------------------------------------------------------------------------
# PASSO 4 — fórmulas dos KPIs / indicadores / avisos (nativas)
# ---------------------------------------------------------------------------
fin = sheets["Financeiro"]
fin["M6"].value = '=SUMIF(FIN_Status,"Pago",FIN_Valor)'
fin["M7"].value = '=SUMIF(FIN_Status,"Pendente",FIN_Valor)'
fin["M8"].value = "=M6-SUM(DESP_Valor)"
fin["M9"].value = "=IFERROR(M7/(M6+M7),0)"
fin["M9"].number_format = "0.0%"
fin["M11"].value = ('=SUMIFS(FIN_Valor,FIN_Status,"Pago",'
                    'FIN_Data,">="&DATE(YEAR(TODAY()),1,1),'
                    'FIN_Data,"<="&DATE(YEAR(TODAY()),12,31))')
fin["M10"].value = ('=IF(MEI_Faturamento>=72900,"Crítico (90%)",'
                    'IF(MEI_Faturamento>=56700,"Atenção (70%)",'
                    '"Dentro do limite"))')
fin["M10"].alignment = ALIGN("left", "center")
fin.column_dimensions["M"].width = 22
for ref in ("M6","M7","M8"):
    fin[ref].number_format = 'R$ #,##0.00'
# rótulo + célula faturamento anual (linha 11)
fin["L11"].value = "Faturamento anual"
fin["L11"].font = mkfont("kpi_label", "3B3B3B")
fin["L11"].alignment = ALIGN("left","center")
fin["L11"].fill = PatternFill(fill_type="solid", fgColor=argb(PAL["cinza_claro"]))
fin["M11"].number_format = 'R$ #,##0.00'
fin["M11"].font = mkfont("kpi_valor_pequeno", "000000")
fin["M11"].alignment = ALIGN("right","center")
fin["M11"].fill = PatternFill(fill_type="solid", fgColor=argb(PAL["cinza_claro"]))

menu = sheets["Menu Inicial"]
menu["J2"].number_format = 'd "de" mmmm "de" yyyy'
menu["B7"].value = '=COUNTIF(PAC_Status,"Ativo")'
menu["C7"].value = ('=SUMIFS(FIN_Valor,FIN_Status,"Pago",'
                    'FIN_Data,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
                    'FIN_Data,"<="&EOMONTH(TODAY(),0))')
menu["C7"].number_format = 'R$ #,##0'
menu["D7"].value = '=SUMIF(FIN_Status,"Pendente",FIN_Valor)'
menu["D7"].number_format = 'R$ #,##0'
menu["F7"].value = ('=IF(SUMPRODUCT((CAL_LogData<>"")*(CAL_LogData>=TODAY()))=0,"—",'
                    'TEXT(MIN(\'Calendário\'!$AB$100:$AB$160),"dd/mm"))')
menu["G7"].value = '=COUNTIF(CAL_LogStatus,"Faltou")'
menu["H7"].value = '=SUMPRODUCT((ANIV_Nasc<>"")*(MONTH(ANIV_Nasc)=MONTH(TODAY())))'

# avisos (texto via fórmula). Override merges/células do bloco lateral.
for mrange in ("J6:K6","J7:K14","J15:K20"):
    try:
        menu.unmerge_cells(mrange)
    except Exception:
        pass
aviso_blocks = [
    ("J6","J6:K7","AVISO_SemSessao",
     '=TEXT(SUMPRODUCT((PAC_UltimaSessao<>"")*(PAC_UltimaSessao<(TODAY()-30))),"0")&" paciente(s) sem sessão há 30+ dias"'),
    ("J9","J9:K10","AVISO_MEI",
     '=IF(MEI_Faturamento>=72900,"Crítico: ",IF(MEI_Faturamento>=56700,"Atenção: ",""))&"Faturamento R$ "&FIXED(MEI_Faturamento,0)&" / teto MEI R$ 81.000"'),
    ("J11","J11:K12","AVISO_Pendentes",
     '=TEXT(COUNTIF(FIN_Status,"Pendente"),"0")&" pagamento(s) pendente(s) — R$ "&FIXED(SUMIF(FIN_Status,"Pendente",FIN_Valor),2)'),
    ("J13","J13:K14","AVISO_Aniversariantes",
     '=TEXT(SUMPRODUCT((ANIV_Nasc<>"")*(TEXT(ANIV_Nasc,"MM-DD")>=TEXT(TODAY(),"MM-DD"))*(TEXT(ANIV_Nasc,"MM-DD")<=TEXT(TODAY()+7,"MM-DD"))),"0")&" aniversariante(s) nos próximos 7 dias"'),
]
_aviso_line = side("thin", PAL["line"])
for ref, mrange, nm, formula in aviso_blocks:
    c = menu[ref]
    c.value = formula
    c.font = mkfont("aviso_texto", "3B3B3B")
    c.alignment = ALIGN("left","center",wrap=True)
    c.fill = PatternFill(fill_type=None)            # sem bloco cinza (redesign)
    c.border = Border(bottom=_aviso_line)           # separador hairline entre avisos
    # marcador vinho discreto na coluna I (acento, não bloco)
    mk = menu["I" + ref[1:]]
    mk.value = "•"
    mk.font = Font(name="Segoe UI Semilight", size=11, color=argb(PAL["vinho_elegante"]))
    mk.alignment = ALIGN("center", "center")
    try:
        menu.merge_cells(mrange)
    except Exception:
        pass

# --- refinamento visual do Menu (otimização premium) ---
_menu_heights = {1:30,2:56,3:22,4:14,5:22,6:24,7:46,8:10,9:22,10:24,
                 11:54,12:14,13:54,14:12,15:8,16:8,17:8,18:8,19:8,20:8}
for _r, _h in _menu_heights.items():
    menu.row_dimensions[_r].height = _h
# wordmark do produto na faixa de cabeçalho vinho: "PSI PLANNER" em Prata DOURADO
menu["B1"].value = "PSI PLANNER"
menu["B1"].font = Font(name="Prata", size=12, bold=False,
                       italic=False, color=argb(PAL["ouro"]))
menu["B1"].alignment = ALIGN("left", "center")
menu.row_dimensions[1].height = 30
# rótulos dos cards: peso fino, cinza médio (hierarquia editorial)
for _ref in ("B6","C6","D6","F6","G6","H6"):
    menu[_ref].font = mkfont("kpi_label", "A6A6A6")
# acento vinho fino SÓ no card crítico (Valor pendente) — cor é sinal
_vinho_side = side("thin", PAL["vinho_elegante"])
_ex = menu["D7"].border
menu["D7"].border = Border(top=_ex.top, left=_ex.left, right=_ex.right, bottom=_vinho_side)
# data do dia recuada (detalhe, não protagonista)
menu["J2"].font = mkfont("data_auto", "A6A6A6")
# título AVISOS em vinho discreto
menu["J5"].font = mkfont("aviso_titulo", "5A1E2A")
# limpa resíduo de preenchimento/borda abaixo do painel (mescla antiga J15:K20)
_no_fill = PatternFill(fill_type=None)
_no_border = Border()
for _row in range(15, 21):
    for _col in ("I", "J", "K", "L"):
        _c = menu[f"{_col}{_row}"]
        _c.fill = _no_fill
        _c.border = _no_border

# Calendário counters
cal = sheets["Calendário"]
cal["K5"].value = "=SUMPRODUCT((CAL_LogData=TODAY())*1)"
cal["K6"].value = '=IFERROR(COUNTIF(CAL_LogStatus,"Cancelado")/COUNTA(CAL_LogStatus),0)'
cal["K6"].number_format = "0.0%"
cal["K7"].value = '=IFERROR(COUNTIF(CAL_LogStatus,"Faltou")/COUNTA(CAL_LogStatus),0)'
cal["K7"].number_format = "0.0%"

# Aniversariantes contador
aniv = sheets["Aniversariantes"]
aniv["K5"].value = '=SUMPRODUCT((ANIV_Nasc<>"")*(MONTH(ANIV_Nasc)=MONTH(TODAY())))'

# Relatórios
rel = sheets["Relatórios"]
rel["C6"].value = "='Menu Inicial'!C7"; rel["C6"].number_format = 'R$ #,##0.00'
rel["C7"].value = "=MEI_Faturamento"; rel["C7"].number_format = 'R$ #,##0.00'
rel["C8"].value = "=COUNTA(CAL_LogStatus)"
rel["F6"].value = '=COUNTIF(PAC_Status,"Ativo")'
rel["F7"].value = "=CAL_TaxaCancelamento"; rel["F7"].number_format = "0.0%"
rel["F8"].value = "=FIN_Inadimplencia"; rel["F8"].number_format = "0.0%"
rel["I6"].value = '=SUMIF(FIN_Forma,"Convênio",FIN_Valor)'; rel["I6"].number_format = 'R$ #,##0.00'
rel["I7"].value = ('=IFERROR((SUMIFS(FIN_Valor,FIN_Status,"Pago",FIN_Data,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),FIN_Data,"<="&EOMONTH(TODAY(),0))'
                   '-SUMIFS(FIN_Valor,FIN_Status,"Pago",FIN_Data,">="&DATE(YEAR(TODAY()),MONTH(TODAY())-1,1),FIN_Data,"<="&EOMONTH(TODAY(),-1)))'
                   '/SUMIFS(FIN_Valor,FIN_Status,"Pago",FIN_Data,">="&DATE(YEAR(TODAY()),MONTH(TODAY())-1,1),FIN_Data,"<="&EOMONTH(TODAY(),-1)),0)')
rel["I7"].number_format = "0.0%"

# ---------------------------------------------------------------------------
# PASSO 5 — dados auxiliares para gráficos (colunas AB+; ocultas)
# ---------------------------------------------------------------------------
MESES = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

def hide_aux(ws, first="AB", last="AP"):
    from openpyxl.utils import column_index_from_string, get_column_letter
    a = column_index_from_string(first); b = column_index_from_string(last)
    for i in range(a, b+1):
        ws.column_dimensions[get_column_letter(i)].hidden = True

# Financeiro aux
fa = sheets["Financeiro"]
fa["AB1"] = "Mês"; fa["AC1"] = "Receita"
for i, m in enumerate(MESES):
    r = 2 + i
    fa[f"AB{r}"] = m
    fa[f"AC{r}"] = (f'=SUMIFS(FIN_Valor,FIN_Status,"Pago",'
                    f'FIN_Data,">="&DATE(YEAR(TODAY()),{i+1},1),'
                    f'FIN_Data,"<="&EOMONTH(DATE(YEAR(TODAY()),{i+1},1),0))')
# Receita x Despesa
fa["AE1"]="Tipo"; fa["AF1"]="Valor"
fa["AE2"]="Receita"; fa["AF2"]="=M6"
fa["AE3"]="Despesa"; fa["AF3"]="=SUM(DESP_Valor)"
# Pago x Pendente
fa["AH1"]="Situação"; fa["AI1"]="Valor"
fa["AH2"]="Pago"; fa["AI2"]="=M6"
fa["AH3"]="Pendente"; fa["AI3"]="=M7"
# Evolução acumulada
fa["AK1"]="Mês"; fa["AL1"]="Acumulado"
for i in range(12):
    r = 2 + i
    fa[f"AK{r}"] = MESES[i]
    if i == 0:
        fa[f"AL{r}"] = "=AC2"
    else:
        fa[f"AL{r}"] = f"=AL{r-1}+AC{2+i}"
# Distribuição forma de pagamento
formas = ["Pix","Dinheiro","Cartão","Convênio","Boleto"]
fa["AN1"]="Forma"; fa["AO1"]="Valor"
for i, fp in enumerate(formas):
    r = 2 + i
    fa[f"AN{r}"] = fp
    fa[f"AO{r}"] = f'=SUMIF(FIN_Forma,"{fp}",FIN_Valor)'
hide_aux(fa)

# Relatórios aux
ra = sheets["Relatórios"]
ra["AB1"]="Mês"; ra["AC1"]="Receita"
for i, m in enumerate(MESES):
    r = 2 + i
    ra[f"AB{r}"] = m
    ra[f"AC{r}"] = (f'=SUMIFS(FIN_Valor,FIN_Status,"Pago",'
                    f'FIN_Data,">="&DATE(YEAR(TODAY()),{i+1},1),'
                    f'FIN_Data,"<="&EOMONTH(DATE(YEAR(TODAY()),{i+1},1),0))')
ra["AE1"]="Mês"; ra["AF1"]="Atend."
for i in range(6):
    r = 2 + i
    ra[f"AE{r}"] = MESES[i]
    ra[f"AF{r}"] = (f'=SUMPRODUCT((MONTH(CAL_LogData)={i+1})*(CAL_LogData<>""))')
conv = ["Particular","Unimed","Bradesco Saúde","SulAmérica","Convênio"]
ra["AH1"]="Convênio"; ra["AI1"]="Valor"
for i, cv in enumerate(conv):
    r = 2 + i
    ra[f"AH{r}"] = cv
    if cv == "Convênio":
        ra[f"AI{r}"] = '=SUMIF(FIN_Forma,"Convênio",FIN_Valor)'
    else:
        ra[f"AI{r}"] = f'=SUMIFS(FIN_Valor,FIN_Paciente,"*")*0+SUMPRODUCT((PAC_Convenio="{cv}")*PAC_ValorSessao)'
hide_aux(ra)

# Aniversariantes aux (por mês)
an = sheets["Aniversariantes"]
an["AB1"]="Mês"; an["AC1"]="Qtd"
for i, m in enumerate(MESES):
    r = 2 + i
    an[f"AB{r}"] = m
    an[f"AC{r}"] = f'=SUMPRODUCT((ANIV_Nasc<>"")*(MONTH(ANIV_Nasc)={i+1}))'
hide_aux(an)
hide_aux(cal)  # esconde coluna auxiliar AB (próxima consulta)

# ---------------------------------------------------------------------------
# PASSO 6 — validações de dados (dropdowns)
# ---------------------------------------------------------------------------
def add_dv(ws, rng, items):
    dv = DataValidation(type="list", formula1='"' + ",".join(items) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)

add_dv(fin, "D7:D56", ["Sessão presencial","Sessão online","Supervisão","Avaliação"])
add_dv(fin, "E7:E56", ["Pix","Dinheiro","Cartão","Convênio","Boleto"])
add_dv(fin, "G7:G56", ["Pago","Pendente","Lembrete","Cancelado"])
add_dv(fin, "H7:H56", ["Emitida","Não emitida"])
add_dv(fin, "C61:C74", ["Aluguel","Supervisão","Marketing","Plataforma online","Materiais","Internet","Outros"])
add_dv(fin, "F61:F74", ["Pix","Dinheiro","Cartão","Convênio","Boleto"])
pac = sheets["Pacientes"]
add_dv(pac, "J7:J56", ["Particular","Unimed","Bradesco Saúde","SulAmérica","Amil"])
add_dv(pac, "M7:M56", ["Semanal","Quinzenal","Mensal"])
add_dv(pac, "O7:O56", ["Ativo","Pausado","Alta","Inativo"])
add_dv(aniv, "G7:G56", ["Ativo","Pausado","Alta","Inativo"])
add_dv(aniv, "H7:H56", ["Sim","Não"])
# Calendário: status com Remarcado/Atendido suportados nos pastéis
add_dv(cal, "C6:H21", ["Confirmado","Atendido","Realizado","Remarcado","Cancelado","Faltou"])
# Configurações: toggle de exibição de convênio
cfg = sheets["Configurações"]
add_dv(cfg, "C18", ["Sim","Não"])
# Configurações: seletor de tema (APARÊNCIA) — alimenta mod_Tema/AplicarTema
add_dv(cfg, "C21", ["Claro","Escuro"])

# ---------------------------------------------------------------------------
# PASSO 7 — formatação condicional
# ---------------------------------------------------------------------------
def cf_fill(fill_key):
    c = argb(PAL[fill_key])
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def cf_font(font_hex):
    return Font(color=argb(font_hex)) if font_hex else None

def contains_rule(ws, rng, anchor, text, fill_key, font_hex):
    formula = [f'ISNUMBER(SEARCH("{text}",{anchor}))']
    ws.conditional_formatting.add(rng, FormulaRule(formula=formula, fill=cf_fill(fill_key), font=cf_font(font_hex)))

# Financeiro status — pastéis explícitos (verde=pago, amarelo=lembrete, vermelho=pendente/cancelado)
contains_rule(fin, "G7:G56", "G7", "Pago", "pastel_green", "1F3B1F")
contains_rule(fin, "G7:G56", "G7", "Pendente", "pastel_red", "3A1018")
contains_rule(fin, "G7:G56", "G7", "Lembrete", "pastel_yellow", "5A4A1E")
contains_rule(fin, "G7:G56", "G7", "Cancelado", "pastel_red", "3A1018")
# MEI indicador (M10) por faixa — tints
fin.conditional_formatting.add("M10", FormulaRule(formula=["MEI_Faturamento>=72900"], fill=cf_fill("alert"), font=cf_font("3A1018"), stopIfTrue=True))
fin.conditional_formatting.add("M10", FormulaRule(formula=["MEI_Faturamento>=56700"], fill=cf_fill("wait"), font=cf_font("000000"), stopIfTrue=True))
# % inadimplência (M9): vinho só acima de 15% (sinal de severidade)
fin.conditional_formatting.add("M9", FormulaRule(formula=["M9>0.15"], font=cf_font("5A1E2A")))
# Pacientes status — pastéis (verde ativo, vermelho inativo)
contains_rule(pac, "O7:O56", "O7", "Ativo", "pastel_green", "1F3B1F")
contains_rule(pac, "O7:O56", "O7", "Pausado", "pastel_yellow", "5A4A1E")
contains_rule(pac, "O7:O56", "O7", "Alta", "surface_deep", "000000")
contains_rule(pac, "O7:O56", "O7", "Inativo", "pastel_red", "3A1018")
# Calendário status (na grade) — pastéis (verde confirmado/atendido, vermelho cancelado/faltou, amarelo remarcado)
contains_rule(cal, "C6:H21", "C6", "Cancelado", "pastel_red", "3A1018")
contains_rule(cal, "C6:H21", "C6", "Faltou", "pastel_red", "3A1018")
contains_rule(cal, "C6:H21", "C6", "Confirmado", "pastel_green", "1F3B1F")
contains_rule(cal, "C6:H21", "C6", "Atendido", "pastel_green", "1F3B1F")
contains_rule(cal, "C6:H21", "C6", "Realizado", "pastel_green", "1F3B1F")
contains_rule(cal, "C6:H21", "C6", "Remarcado", "pastel_yellow", "5A4A1E")
# Aniversariantes: semana = barra vinho à ESQUERDA do nome (acento, não bloco)
_wk = 'AND($C7<>"",TEXT($C7,"MM-DD")>=TEXT(TODAY(),"MM-DD"),TEXT($C7,"MM-DD")<=TEXT(TODAY()+7,"MM-DD"))'
aniv.conditional_formatting.add("B7:B56", FormulaRule(
    formula=[_wk],
    border=Border(left=Side(style="medium", color=argb(PAL["vinho_elegante"]))),
    font=Font(color=argb(PAL["vinho_elegante"]))))
# Aniversariantes — Status do paciente (G): verde ativo, vermelho inativo
contains_rule(aniv, "G7:G56", "G7", "Ativo", "pastel_green", "1F3B1F")
contains_rule(aniv, "G7:G56", "G7", "Pausado", "pastel_yellow", "5A4A1E")
contains_rule(aniv, "G7:G56", "G7", "Alta", "surface_deep", "000000")
contains_rule(aniv, "G7:G56", "G7", "Inativo", "pastel_red", "3A1018")
# Aniversariantes — Mensagem enviada (H): verde Sim, vermelho Não
contains_rule(aniv, "H7:H56", "H7", "Sim", "pastel_green", "1F3B1F")
contains_rule(aniv, "H7:H56", "H7", "Não", "pastel_red", "3A1018")
# Relatórios: taxas em vinho acima do limite (sinal)
rel.conditional_formatting.add("F8", FormulaRule(formula=["F8>0.15"], font=cf_font("5A1E2A")))
rel.conditional_formatting.add("F7", FormulaRule(formula=["F7>0.1"], font=cf_font("5A1E2A")))

# ---------------------------------------------------------------------------
# PASSO 8 — gráficos (monocromáticos, 1 série/destaque vinho)
# ---------------------------------------------------------------------------
# paleta CATEGÓRICA da identidade — distinguível em claro E escuro
VINHO = "5A1E2A"; OURO = "C9A86A"; SAGE = "6E8E6E"; TERRA = "C0805E"; SLATE = "7E8AA8"
CAT = [VINHO, OURO, SAGE, TERRA, SLATE]   # 5 categorias bem distintas
CINZA_M = "9A9A9A"

def set_series_color(s, hexcolor):
    s.graphicalProperties.solidFill = hexcolor
    s.graphicalProperties.line.solidFill = hexcolor

# --- tipografia de gráfico na identidade (Prata título + Inter eixos) ---
def _chart_title(text):
    cp = CharacterProperties(latin=DrawFont(typeface="Prata"), sz=1100, b=False, i=False, solidFill="3B3B3B")
    para = Paragraph(pPr=ParagraphProperties(defRPr=cp), r=[RegularTextRun(rPr=cp, t=text)])
    t = Title(tx=Text(rich=RichText(p=[para])))
    t.overlay = False   # título não cobre o gráfico/rosca
    return t

def _axis_style(ch):
    cp = CharacterProperties(latin=DrawFont(typeface="Inter"), sz=800, solidFill="3B3B3B")
    txpr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    if ch.x_axis is not None:
        ch.x_axis.txPr = txpr
    if ch.y_axis is not None:
        ch.y_axis.txPr = txpr

def _chart_frame(ch):
    ch.graphical_properties = GraphicalProperties(ln=LineProperties(solidFill="E8E8E8"))

def add_line(ws, title, anchor, data_ref, cats_ref, color=VINHO):
    ch = LineChart(); ch.title = _chart_title(title); ch.style = 2
    ch.height = 6; ch.width = 9
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    for s in ch.series:
        set_series_color(s, color)
        s.smooth = False
    ch.legend = None
    ch.y_axis.majorGridlines = None
    ch.x_axis.majorGridlines = None
    ch.y_axis.delete = False
    ch.x_axis.delete = False
    ch.y_axis.scaling.min = 0   # eixo de valor a partir de 0
    _axis_style(ch); _chart_frame(ch)
    ws.add_chart(ch, anchor)

def add_bar_h(ws, title, anchor, data_ref, cats_ref, color=VINHO, point_colors=None):
    ch = BarChart(); ch.type = "bar"; ch.title = _chart_title(title); ch.style = 2
    ch.height = 6; ch.width = 9
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    if point_colors:   # cor por barra (ex.: Receita verde x Despesa vinho)
        ch.varyColors = True            # honra cor por ponto
        s = ch.series[0]
        s.data_points = []
        for i, col in enumerate(point_colors):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = col
            dp.graphicalProperties.line.solidFill = col
            s.data_points.append(dp)
    else:
        for s in ch.series:
            set_series_color(s, color)
    ch.legend = None
    ch.y_axis.majorGridlines = None
    ch.x_axis.majorGridlines = None
    ch.y_axis.delete = False
    ch.x_axis.delete = False
    ch.y_axis.scaling.min = 0   # eixo de VALOR (y_axis no openpyxl, mesmo em barra h.) começa em 0
    _axis_style(ch); _chart_frame(ch)
    ws.add_chart(ch, anchor)

def add_doughnut(ws, title, anchor, data_ref, cats_ref, colors, height=6, width=9):
    ch = DoughnutChart(); ch.title = _chart_title(title); ch.holeSize = 55
    ch.height = height; ch.width = width
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    s = ch.series[0]
    s.data_points = []
    for i, col in enumerate(colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = col
        dp.graphicalProperties.line.solidFill = "FAFAF8"   # borda fina separa as fatias
        s.data_points.append(dp)
    _chart_frame(ch)
    ws.add_chart(ch, anchor)

# Financeiro: 5 gráficos
add_line(fin, "Receita mensal", "B78", Reference(fin, min_col=29, min_row=1, max_row=13), Reference(fin, min_col=28, min_row=2, max_row=13), VINHO)
add_bar_h(fin, "Receita x Despesa", "F78", Reference(fin, min_col=32, min_row=1, max_row=3), Reference(fin, min_col=31, min_row=2, max_row=3), point_colors=[SAGE, VINHO])
add_doughnut(fin, "Pago x Pendente", "J78", Reference(fin, min_col=35, min_row=1, max_row=3), Reference(fin, min_col=34, min_row=2, max_row=3), [SAGE, VINHO])
add_line(fin, "Evolução acumulada", "B96", Reference(fin, min_col=38, min_row=1, max_row=13), Reference(fin, min_col=37, min_row=2, max_row=13), VINHO)
add_doughnut(fin, "Distribuição por forma", "F96", Reference(fin, min_col=41, min_row=1, max_row=6), Reference(fin, min_col=40, min_row=2, max_row=6), CAT)

# Relatórios: 3 gráficos
add_line(rel, "Evolução da receita", "B16", Reference(rel, min_col=29, min_row=1, max_row=13), Reference(rel, min_col=28, min_row=2, max_row=13), VINHO)
add_bar_h(rel, "Atendimentos por mês", "E16", Reference(rel, min_col=32, min_row=1, max_row=7), Reference(rel, min_col=31, min_row=2, max_row=7), color=VINHO)
add_doughnut(rel, "Receita por convênio", "H16", Reference(rel, min_col=35, min_row=1, max_row=6), Reference(rel, min_col=34, min_row=2, max_row=6), CAT)

# Aniversariantes: 1 doughnut colorido (1 cor por mês, mantendo o vinho como destaque do mês atual)
ANIV_COLORS = [
    "E07A5F",  # Jan - terracota
    "F2A65A",  # Fev - âmbar
    "F2C14E",  # Mar - mostarda
    "B8C77A",  # Abr - oliva
    "7AB89E",  # Mai - sálvia
    "5C9EAD",  # Jun - azul-petróleo claro
    "4B6EAF",  # Jul - azul
    "8E7DBE",  # Ago - lavanda
    "B25A85",  # Set - rosa-vinho
    "5A1E2A",  # Out - vinho (cor da marca)
    "C26F5A",  # Nov - coral
    "8D9F3F",  # Dez - oliva escuro
]
# Destaca o mês atual em vinho (mantém o resto da paleta colorida)
import datetime as _dt
_this_month_idx = _dt.date.today().month - 1
_aniv_palette = list(ANIV_COLORS)
_aniv_palette[_this_month_idx] = "5A1E2A"
add_doughnut(aniv, "Aniversariantes por mês", "J8", Reference(aniv, min_col=29, min_row=1, max_row=13), Reference(aniv, min_col=28, min_row=2, max_row=13),
             _aniv_palette, height=8, width=13)

# ---------------------------------------------------------------------------
# PASSO 9 — Configurações: ajustar profissional/registro
# ---------------------------------------------------------------------------
cfg = sheets["Configurações"]
cfg["C7"].value = "Guilherme Quintino"
cfg["C8"].value = "CRP 09/12345"
cfg["C9"].value = "(62) 99000-1010"
cfg["C10"].value = "contato@clinicavita.com.br"
cfg["C13"].value = 81000
cfg["C13"].number_format = 'R$ #,##0'
cfg["C14"].value = 180
cfg["C14"].number_format = 'R$ #,##0.00'

# --- respiro visual: alturas de linha (otimização premium) ---
for _r in range(7, 57):
    fin.row_dimensions[_r].height = 20
    pac.row_dimensions[_r].height = 20
    aniv.row_dimensions[_r].height = 20
for _r in range(61, 75):
    fin.row_dimensions[_r].height = 20
for _r in (6,):  # cabeçalhos de tabela mais respirados
    fin.row_dimensions[_r].height = 28
    pac.row_dimensions[_r].height = 28
    aniv.row_dimensions[_r].height = 28
# grade do Calendário mais alta (evita corte de "Tipo · Status")
for _r in range(6, 22):
    cal.row_dimensions[_r].height = 34

# congelar painéis nas tabelas
fin.freeze_panes = "B7"
pac.freeze_panes = "B7"
aniv.freeze_panes = "B7"
# esconder gridlines E cabeçalhos de linha/coluna (visual de app, não planilha)
for ws in sheets.values():
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.zoomScale = 100

# ---------------------------------------------------------------------------
# PASSO 10 — redesign: réguas de cabeçalho, header do calendário, cor de texto
# ---------------------------------------------------------------------------
_accent = side("thin", PAL["vinho_elegante"])
_lhair = side("thin", PAL["line"])
def _header_accent(ws, rng):
    for _row in ws[rng]:
        for _c in _row:
            _ex = _c.border
            _c.border = Border(top=_ex.top, left=_ex.left, right=_ex.right, bottom=_accent)
for _ws, _rng in [(fin, "B6:J6"), (fin, "B60:F60"), (pac, "B6:T6"), (aniv, "B6:H6")]:
    _header_accent(_ws, _rng)
# cabeçalho do Calendário: ink sobre surface (não barra escura) + régua vinho
for _col in "BCDEFGH":
    _cc = cal[f"{_col}5"]
    _cc.fill = PatternFill(fill_type="solid", fgColor=argb(PAL["cinza_claro"]))
    _cc.font = mkfont("cabecalho_tabela", "000000")
_header_accent(cal, "B5:H5")
# indicadores Financeiro: hairline separando cada par (sem caixa)
for _r in range(6, 11):
    for _col in ("L", "M"):
        _cc = fin[f"{_col}{_r}"]
        _ex = _cc.border
        _cc.border = Border(top=_ex.top, left=_ex.left, right=_ex.right, bottom=_lhair)

# normalização de cor de texto -> 2 níveis: ink (#000000) + mute (#3B3B3B)
_REMAP = {"FF000000": "FF000000", "FF3B3B3B": "FF3B3B3B", "FFA6A6A6": "FF3B3B3B"}
for _ws in sheets.values():
    for _row in _ws.iter_rows():
        for _c in _row:
            _f = _c.font
            if _f and _f.color is not None and isinstance(getattr(_f.color, "rgb", None), str) and _f.color.rgb in _REMAP:
                _c.font = Font(name=_f.name, size=_f.size, bold=_f.bold, italic=_f.italic,
                               color=_REMAP[_f.color.rgb])

# ---------------------------------------------------------------------------
# PASSO 11 — cores de aba, prontidão para impressão, campos de input
# ---------------------------------------------------------------------------
# cores de aba: Menu em vinho (home), demais em mute — barra inferior coesa
menu.sheet_properties.tabColor = "5A1E2A"
for _nm in ("Financeiro", "Pacientes", "Calendário", "Aniversariantes", "Relatórios", "Configurações"):
    sheets[_nm].sheet_properties.tabColor = "3B3B3B"

def _setup_print(ws, area, landscape=True, repeat=None):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    # NÃO definir ws.print_area: em .xlsm a presença de _xlnm.Print_Area faz o
    # Excel paginar contra o driver de impressora ao ABRIR (pode travar). O fit-
    # to-width/orientação sobrevivem; a impressão usa o range visível.
    if repeat:
        ws.print_title_rows = repeat

_setup_print(menu, "A1:L21", landscape=True)
_setup_print(fin, "A1:M74", landscape=True, repeat="1:6")
_setup_print(pac, "A1:T56", landscape=True, repeat="1:6")
_setup_print(cal, "A1:K22", landscape=True)
_setup_print(aniv, "A1:K20", landscape=True, repeat="1:6")
_setup_print(rel, "A1:I30", landscape=True)
_setup_print(cfg, "A1:D14", landscape=False)

# campos de input em Configurações: linha de base (affordance de edição)
_input_side = side("thin", PAL["mute"])
for _r in (6, 7, 8, 9, 10, 13, 14):
    _cc = cfg[f"C{_r}"]
    _cc.border = Border(bottom=_input_side)

# ---------------------------------------------------------------------------
# PASSO 12 — polimento final: rodapé de impressão, abertura limpa, log oculto
# ---------------------------------------------------------------------------
for _ws in sheets.values():
    _ws.oddFooter.left.text = "Psi Planner"   # marca do produto (rodapé impresso)
    _ws.oddFooter.left.size = 8
    _ws.oddFooter.left.color = "3B3B3B"
    _ws.oddFooter.center.text = "Clínica Vita"
    _ws.oddFooter.center.size = 8
    _ws.oddFooter.center.color = "3B3B3B"
    _ws.oddFooter.right.text = "&P / &N"
    _ws.oddFooter.right.size = 8
    _ws.oddFooter.right.color = "3B3B3B"
    # abre cada aba com a célula A1 selecionada (sem âncora em célula aleatória)
    try:
        _ws.sheet_view.selection[0].activeCell = "A1"
        _ws.sheet_view.selection[0].sqref = "A1"
    except Exception:
        pass
# Calendário: ocultar a base de log da agenda (alimenta contadores, fica fora da vista)
cal["B97"].value = None
for _r in range(96, 161):
    cal.row_dimensions[_r].hidden = True

# propriedades do documento (identidade do produto)
wb.properties.title = PRODUTO
wb.properties.subject = "Planejador financeiro-clínico para psicólogos e terapeutas"
wb.properties.creator = PRODUTO
wb.properties.keywords = "psicologia, financeiro, MEI, agenda, clínica"
wb.properties.description = ("Psi Planner — controle financeiro, agenda e indicadores "
                             "para profissionais de saúde mental.")
wb.properties.category = "Planilha financeira"

# ---------------------------------------------------------------------------
# PASSO 13 — TAREFA 2: superfície off-white (base do tema CLARO)
# Pinta o retângulo visível de cada aba com FAFAF8 SOMENTE onde não há fill,
# preservando cards/cabeçalhos (F4F4F4) e demais cores. Define a "surface base"
# do tema claro (substitui o branco puro do fundo). Os mesmos retângulos são
# usados por mod_Tema.AplicarTema para repintar Claro<->Escuro.
# ---------------------------------------------------------------------------
from openpyxl.utils import range_boundaries
PAINT_RECTS = {
    "Menu Inicial":    "A1:L24",
    "Financeiro":      "A1:N118",
    "Pacientes":       "A1:T58",
    "Calendário":      "A1:K25",
    "Aniversariantes": "A1:Q24",
    "Relatórios":      "A1:O34",
    "Configurações":   "A1:D26",
}
_OFFWHITE = argb(PAL["off_white"])  # FAFAF8
for _nm, _rect in PAINT_RECTS.items():
    _ws = sheets[_nm]
    _minc, _minr, _maxc, _maxr = range_boundaries(_rect)
    for _r in range(_minr, _maxr + 1):
        for _c in range(_minc, _maxc + 1):
            _cell = _ws.cell(row=_r, column=_c)
            if _cell.fill is None or _cell.fill.fill_type is None:
                _cell.fill = PatternFill(fill_type="solid", fgColor=_OFFWHITE)

# ---------------------------------------------------------------------------
# PASSO 14 — IDENTIDADE: faixa de cabeçalho vinho + filete dourado + monograma
# Faixa (linha 1) em vinho com o logo dourado, coerente nos DOIS temas (o
# mod_Tema EXCLUI a linha 1 do remap, então a faixa permanece vinho/ouro).
# Filete dourado (borda superior da linha 2) separa a faixa do conteúdo.
# ---------------------------------------------------------------------------
from openpyxl.drawing.image import Image as XLImage
_BAND = argb(PAL["vinho_elegante"])           # faixa vinho 5A1E2A (constante)
_filete = side("thin", PAL["ouro_escuro"])     # filete dourado (tematizado p/ ouro no escuro)
MONO_ANCHOR = {
    "Menu Inicial": "K1", "Financeiro": "M1", "Pacientes": "S1",
    "Calendário": "J1", "Aniversariantes": "P1", "Relatórios": "N1",
    "Configurações": "C1",
}
MONO_PATH = BUILD + r"\assets\psi_monograma_ouro.png"
for _nm, _rect in PAINT_RECTS.items():
    _ws = sheets[_nm]
    _minc, _minr, _maxc, _maxr = range_boundaries(_rect)
    _ws.row_dimensions[1].height = 30
    for _col in range(_minc, _maxc + 1):
        _ws.cell(row=1, column=_col).fill = PatternFill(fill_type="solid", fgColor=_BAND)
        _c2 = _ws.cell(row=2, column=_col)
        _ex = _c2.border
        _c2.border = Border(top=_filete, left=_ex.left, right=_ex.right, bottom=_ex.bottom)
    try:
        _img = XLImage(MONO_PATH)
        _img.width = 26
        _img.height = 26
        _ws.add_image(_img, MONO_ANCHOR.get(_nm, "A1"))
    except Exception as _e:
        print("logo warn (%s): %r" % (_nm, _e))

wb.save(OUT_XLSX)
print("OK ->", OUT_XLSX)
