# -*- coding: utf-8 -*-
"""
Bateria de testes do Psi Planner.xlsm.
PARTE A (openpyxl): estrutura, named ranges, formatos, freeze, impressão,
                    rodapé, cores de aba, colunas ocultas, CF, branding, props.
PARTE B (COM):      células de erro, navegação, subs de cálculo, KPIs ao vivo,
                    gráficos, GerarRecibo, RegistrarConsulta, export PDF, procs.
PARTE C:            contraste WCAG da paleta.
Saída: lista de OK/FAIL + RESULTADO final; exit code != 0 se houver falha.
"""
import os, sys, time, glob
import openpyxl
import win32com.client as win32

FINAL = r"D:\Vittor\Projetos\Saluti\PsiPlanner\modelo2\Psi Planner 2.xlsm"
DELIV = r"D:\Vittor\Projetos\Saluti\PsiPlanner\modelo2"

SHEETS = ["Menu Inicial", "Financeiro", "Pacientes", "Calendário",
          "Aniversariantes", "Relatórios", "Configurações"]

# valores esperados (calculados a partir do seed; hoje = 2026-05-27)
DESPESAS = [99, 110, 600, 300, 150, 120, 80]   # soma = 1459
EXP = {
    "pac_ativos": 7,
    "receita_mes": 870,      # Mai/2026 pagos
    "valor_pendente": 350,   # Diego 200 + Kris 150
    "recebido": 3660,
    "pendente": 350,
    "lucro": 3660 - sum(DESPESAS),   # 101
    "inadimplencia": 350 / (3660 + 350),
    "fat_anual": 3660,
    "faltas": 1,
    "aniversariantes": 3,    # Bruno, Camila, Eduarda (Mai)
}

fails, warns = [], []
def check(ok, msg):
    print(("  OK  " if ok else " FAIL ") + msg)
    if not ok:
        fails.append(msg)
def near(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False

# ==========================================================================
# PARTE A — openpyxl (estrutura estática lida do arquivo salvo)
# ==========================================================================
print("===== PARTE A: ESTRUTURA (openpyxl) =====")
wb = openpyxl.load_workbook(FINAL, data_only=False, keep_vba=True)

print("== A1. Abas e ordem ==")
check(wb.sheetnames == SHEETS, "abas exatas/ordem: %s" % wb.sheetnames)

print("== A2. Propriedades do documento ==")
check((wb.properties.title or "") == "Psi Planner 2",
      "Title do documento = %r" % wb.properties.title)

print("== A3. Branding: eyebrow + sem 'Saluti' em célula alguma ==")
menu = wb["Menu Inicial"]
check((menu["B1"].value or "") == "PSI PLANNER 2", "Menu!B1 eyebrow = %r" % menu["B1"].value)
leftovers = []
for sn in wb.sheetnames:
    for row in wb[sn].iter_rows():
        for c in row:
            if isinstance(c.value, str) and "saluti" in c.value.lower():
                leftovers.append("%s!%s=%r" % (sn, c.coordinate, c.value))
check(not leftovers, "nenhum 'Saluti' residual em células (achados: %s)" % leftovers[:5])

print("== A4. Named ranges essenciais resolvem ==")
need_names = ["KPI_PacientesAtivos", "KPI_ReceitaMes", "KPI_ValorPendente",
              "KPI_ProxConsulta", "KPI_FaltasMes", "KPI_AniversariantesMes",
              "FIN_Valor", "FIN_Status", "FIN_Data", "DESP_Valor", "PAC_Status",
              "PAC_Nome", "PAC_UltimaSessao", "ANIV_Nasc", "ANIV_Nome",
              "CAL_LogData", "CAL_LogStatus", "MEI_Faturamento", "MEI_Indicador",
              "FIN_TotalRecebido", "CFG_NomeClinica", "CFG_LimiteMEI", "CFG_ValorSessao"]
defined = set(wb.defined_names.keys())
miss = [n for n in need_names if n not in defined]
check(not miss, "todos os named ranges presentes (faltando: %s)" % miss)

print("== A5. Formatos numéricos ==")
fin = wb["Financeiro"]; pac = wb["Pacientes"]; cal = wb["Calendário"]
def is_money(fmt):   # openpyxl escapa: "\R\$\ #,##0.00"
    return "R" in (fmt or "") and "$" in (fmt or "")
def is_date(fmt):    # aceita "dd/mm/yyyy" ou builtin locale curto (mm-dd-yy)
    f = (fmt or "").lower()
    return "d" in f and "m" in f and "y" in f
check(is_money(fin["F7"].number_format), "Financeiro F7 moeda (%s)" % fin["F7"].number_format)
check(is_date(fin["B7"].number_format), "Financeiro B7 data (%s)" % fin["B7"].number_format)
check(fin["M9"].number_format == "0.0%", "Financeiro M9 percentual (%s)" % fin["M9"].number_format)
check(is_money(menu["C7"].number_format), "Menu C7 moeda (%s)" % menu["C7"].number_format)
check(is_date(pac["E7"].number_format), "Pacientes E7 data (%s)" % pac["E7"].number_format)
check(is_money(pac["N7"].number_format), "Pacientes N7 moeda (%s)" % pac["N7"].number_format)
check(cal["K6"].number_format == "0.0%", "Calendário K6 percentual (%s)" % cal["K6"].number_format)

print("== A5b. Fontes aplicadas (Golsuka-wordmark/Playfair/Poppins) ==")
check(fin["B2"].font.name == "Playfair Display", "título aba (B2) = %r" % fin["B2"].font.name)
check(fin["B3"].font.name == "Playfair Display" and bool(fin["B3"].font.italic),
      "subtítulo (B3) Playfair itálico = %r it=%s" % (fin["B3"].font.name, fin["B3"].font.italic))
check(fin["C7"].font.name == "Poppins Light", "conteúdo (C7) = %r" % fin["C7"].font.name)
check(menu["B1"].font.name == "Poppins Light", "eyebrow (B1) = %r" % menu["B1"].font.name)
check(menu["B2"].font.name == "Playfair Display", "nome clínica (B2) = %r" % menu["B2"].font.name)

print("== A6. Validações de dados (dropdowns) ==")
def _cell_in_range(cell, rng):
    from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple
    try:
        minc, minr, maxc, maxr = range_boundaries(rng)
        r, c = coordinate_to_tuple(cell)
        return (minr <= r <= maxr) and (minc <= c <= maxc)
    except Exception:
        return False
def dv_covers(ws, cell):
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and any(_cell_in_range(cell, rng) for rng in str(dv.sqref).split()):
            return True
    return False
dv_expect = [("Financeiro", c) for c in ("D7", "E7", "G7", "H7", "C61", "F61")] + \
            [("Pacientes", c) for c in ("J7", "M7", "O7")] + \
            [("Aniversariantes", c) for c in ("G7", "H7")]
for sn, cell in dv_expect:
    check(dv_covers(wb[sn], cell), "validação lista em %s!%s" % (sn, cell))

print("== A7. Formatação condicional (rules por aba) ==")
def cf_count(ws):
    n = 0
    for cf in ws.conditional_formatting:
        rules = getattr(cf, "rules", None)
        if rules is None:
            rules = ws.conditional_formatting[cf]
        n += len(rules)
    return n
cf_expect = {"Financeiro": 7, "Pacientes": 4, "Calendário": 4,
             "Aniversariantes": 3, "Relatórios": 2}
for sn, mn in cf_expect.items():
    got = cf_count(wb[sn])
    check(got >= mn, "%s tem %d CF rules (>= %d)" % (sn, got, mn))

print("== A8. Freeze panes ==")
for sn in ("Financeiro", "Pacientes", "Aniversariantes"):
    check(wb[sn].freeze_panes == "B7", "%s freeze=%s" % (sn, wb[sn].freeze_panes))

print("== A9. Impressão (orientação; print_area omitido de propósito) ==")
# print_area NÃO é definido no .xlsm: em .xlsm a presença de _xlnm.Print_Area faz
# o Excel paginar contra a impressora ao ABRIR, podendo travar a abertura. A
# impressão usa o range visível (oculto não imprime). Conferimos só a orientação.
exp_orient = {"Configurações": "portrait"}
for sn in SHEETS:
    want = exp_orient.get(sn, "landscape")
    check(wb[sn].page_setup.orientation == want,
          "%s orientação=%s (esp. %s)" % (sn, wb[sn].page_setup.orientation, want))
    check(not wb[sn].print_area, "%s sem print_area (evita travar abertura) = %r" % (sn, wb[sn].print_area))

print("== A10. Rodapé com marca do produto ==")
for sn in SHEETS:
    ft = (wb[sn].oddFooter.left.text or "") + (wb[sn].oddFooter.center.text or "")
    check("Psi Planner" in ft, "%s rodapé contém 'Psi Planner'" % sn)

print("== A11. Cores de aba ==")
def tabrgb(ws):
    tc = ws.sheet_properties.tabColor
    return getattr(tc, "rgb", None) if tc is not None else None
check(str(tabrgb(menu)).endswith("5A1E2A"), "Menu tab vinho (%s)" % tabrgb(menu))
for sn in ("Financeiro", "Pacientes", "Calendário", "Aniversariantes", "Relatórios", "Configurações"):
    check(str(tabrgb(wb[sn])).endswith("3B3B3B"), "%s tab mute (%s)" % (sn, tabrgb(wb[sn])))

print("== A12. Colunas auxiliares ocultas + gridlines/headers off ==")
check(fin.column_dimensions["AB"].hidden, "Financeiro AB oculta")
for sn in SHEETS:
    ws = wb[sn]
    check(not ws.sheet_view.showGridLines, "%s gridlines off" % sn)
    check(not ws.sheet_view.showRowColHeaders, "%s headers off" % sn)

# ==========================================================================
# PARTE B — COM (comportamento ao vivo)
# ==========================================================================
print("\n===== PARTE B: COMPORTAMENTO (COM) =====")
xlCellTypeFormulas, xlErrors = -4123, 16
xlTypePDF = 0
ex = win32.gencache.EnsureDispatch("Excel.Application")
ex.Visible = False; ex.DisplayAlerts = False; ex.AutomationSecurity = 1
try:
    wbk = ex.Workbooks.Open(FINAL)
    ex.Calculation = -4105
    ex.CalculateFull(); time.sleep(0.4)
    m = wbk.Worksheets("Menu Inicial"); f = wbk.Worksheets("Financeiro")

    print("== B1. Sem células de erro (todas as abas) ==")
    total_err = 0
    for s in SHEETS:
        ws = wbk.Worksheets(s)
        try:
            errs = ws.UsedRange.SpecialCells(xlCellTypeFormulas, xlErrors)
            addrs = [c.Address(False, False) for ar in errs.Areas for c in ar.Cells]
            total_err += len(addrs)
            if addrs:
                check(False, "%s: %d erro(s) %s" % (s, len(addrs), addrs[:5]))
        except Exception:
            pass
    check(total_err == 0, "total de células com erro = %d" % total_err)

    print("== B2. KPIs ao vivo conferem com valores esperados ==")
    check(near(m.Range("B7").Value, EXP["pac_ativos"]), "pacientes ativos = %s (esp %s)" % (m.Range("B7").Value, EXP["pac_ativos"]))
    check(near(m.Range("C7").Value, EXP["receita_mes"]), "receita mês = %s (esp %s)" % (m.Range("C7").Value, EXP["receita_mes"]))
    check(near(m.Range("D7").Value, EXP["valor_pendente"]), "valor pendente = %s (esp %s)" % (m.Range("D7").Value, EXP["valor_pendente"]))
    check(near(m.Range("G7").Value, EXP["faltas"]), "faltas = %s (esp %s)" % (m.Range("G7").Value, EXP["faltas"]))
    check(near(m.Range("H7").Value, EXP["aniversariantes"]), "aniversariantes = %s (esp %s)" % (m.Range("H7").Value, EXP["aniversariantes"]))
    px = m.Range("F7").Value
    check(px not in (None, "", "—", "-"), "próxima consulta resolvida = %r" % px)
    check(near(f.Range("M6").Value, EXP["recebido"]), "recebido (M6) = %s" % f.Range("M6").Value)
    check(near(f.Range("M7").Value, EXP["pendente"]), "pendente (M7) = %s" % f.Range("M7").Value)
    check(near(f.Range("M8").Value, EXP["lucro"]), "lucro líquido (M8) = %s (esp %s)" % (f.Range("M8").Value, EXP["lucro"]))
    check(near(f.Range("M9").Value, EXP["inadimplencia"], 0.01), "inadimplência (M9) = %s" % f.Range("M9").Value)
    check(near(f.Range("M11").Value, EXP["fat_anual"]), "faturamento anual (M11) = %s" % f.Range("M11").Value)
    check(str(f.Range("M10").Value).strip().lower().startswith("dentro"), "indicador MEI (M10) = %r" % f.Range("M10").Value)

    print("== B3. Navegação (7 macros) ==")
    nav = {"IrFinanceiro": "Financeiro", "IrPacientes": "Pacientes",
           "IrCalendario": "Calendário", "IrAniversariantes": "Aniversariantes",
           "IrRelatorios": "Relatórios", "IrConfiguracoes": "Configurações",
           "IrMenu": "Menu Inicial"}
    for sub, dest in nav.items():
        try:
            ex.Run(sub); ok = (ex.ActiveSheet.Name == dest)
        except Exception:
            ok = False
        check(ok, "%s -> %s" % (sub, dest))
    ex.Run("IrMenu")

    print("== B4. Gráficos presentes e com séries ==")
    chart_expect = {"Financeiro": 5, "Relatórios": 3, "Aniversariantes": 1,
                    "Menu Inicial": 0, "Pacientes": 0, "Calendário": 0, "Configurações": 0}
    for sn, n in chart_expect.items():
        ws = wbk.Worksheets(sn)
        cnt = ws.ChartObjects().Count
        check(cnt == n, "%s tem %d gráfico(s) (esp %d)" % (sn, cnt, n))
        for i in range(1, cnt + 1):
            co = ws.ChartObjects(i)
            try:
                sc = co.Chart.SeriesCollection().Count
                vals = co.Chart.SeriesCollection(1).Values
                ok = sc >= 1 and vals is not None and len(vals) > 0
            except Exception as e:
                ok = False
            check(ok, "  %s gráfico #%d com série/dados" % (sn, i))

    print("== B5. Procedimentos interativos presentes (compila pois macros rodam) ==")
    allcode = ""
    for comp in wbk.VBProject.VBComponents:
        cm = comp.CodeModule
        if cm.CountOfLines > 0:
            allcode += cm.Lines(1, cm.CountOfLines) + "\n"
    for proc in ["GerarRecibo", "ImprimirRecibo", "ExportarRelatorioPDF",
                 "AbrirFormConsulta", "RegistrarConsulta", "RemarcarConsulta",
                 "FinalizarConsulta", "BuscarPaciente", "LimparFiltros",
                 # filtros documentados no spec ("Filtros inteligentes")
                 "FiltrarPorStatus", "FiltrarPorConvenio", "FiltrarPorMesAniversario",
                 "FiltrarPorFrequencia", "FiltrarPorFaturamento"]:
        check(("Sub " + proc) in allcode, "procedimento %s definido" % proc)

    print("== B6. Subs de cálculo não quebram ==")
    for sub in ["AtualizarAvisos", "AtualizarAlertaMEI", "AtualizarIdades", "LimparFiltros"]:
        try:
            ex.Run(sub); check(True, "%s executou" % sub)
        except Exception as e:
            check(False, "%s erro: %r" % (sub, e))

    print("== B7. GerarRecibo (fluxo sucesso, sem MsgBox) ==")
    try:
        f.Activate(); f.Range("C7").Select()
        ex.Run("GerarRecibo")
        rec = wbk.Worksheets("Recibo")
        hdr = str(rec.Range("A1").Value or "")
        pacv = str(rec.Range("B5").Value or "")
        clin = str(wbk.Worksheets("Configurações").Range("C6").Value or "")
        check("SALUTI" not in hdr.upper(), "recibo header sem 'SALUTI' (=%r)" % hdr)
        check(hdr == clin and hdr != "", "recibo header = nome da clínica (%r)" % hdr)
        check(pacv == "Eduarda Faria", "recibo paciente (1ª linha) = %r" % pacv)
        ex.DisplayAlerts = False; rec.Delete()
    except Exception as e:
        check(False, "GerarRecibo falhou: %r" % e)

    print("== B8. RegistrarConsulta (append no log da agenda) ==")
    try:
        cal = wbk.Worksheets("Calendário")
        lin = 100
        while cal.Cells(lin, 1).Value not in (None, ""):
            lin += 1
        ex.Run("RegistrarConsulta", "Teste QA", "01/06/2026", "10:00", "Online", "Confirmado")
        check(str(cal.Cells(lin, 3).Value) == "Teste QA",
              "consulta registrada na linha %d (C=%r)" % (lin, cal.Cells(lin, 3).Value))
    except Exception as e:
        check(False, "RegistrarConsulta falhou: %r" % e)

    print("== B9. Exportar PDF (capacidade ExportAsFixedFormat) ==")
    try:
        outdir = os.path.join(os.path.dirname(FINAL), "Exportados")
        os.makedirs(outdir, exist_ok=True)
        pdf = os.path.join(outdir, "qa_test_export.pdf")
        if os.path.exists(pdf):
            os.remove(pdf)
        rel = wbk.Worksheets("Relatórios"); rel.Activate()
        rel.ExportAsFixedFormat(xlTypePDF, pdf)
        time.sleep(0.6)
        check(os.path.exists(pdf) and os.path.getsize(pdf) > 0, "PDF exportado (%s bytes)" % (os.path.getsize(pdf) if os.path.exists(pdf) else 0))
        if os.path.exists(pdf):
            os.remove(pdf)
    except Exception as e:
        check(False, "Export PDF falhou: %r" % e)

    wbk.Close(SaveChanges=False)
finally:
    ex.Quit()

# limpa pasta Exportados se vazia
try:
    od = os.path.join(DELIV, "Exportados")
    if os.path.isdir(od) and not os.listdir(od):
        os.rmdir(od)
except Exception:
    pass

# ==========================================================================
# PARTE C — contraste WCAG
# ==========================================================================
print("\n===== PARTE C: CONTRASTE WCAG =====")
def lum(hx):
    hx = hx.lstrip("#"); r, g, b = (int(hx[i:i+2], 16) / 255 for i in (0, 2, 4))
    fn = lambda c: c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
    R, G, B = fn(r), fn(g), fn(b); return 0.2126 * R + 0.7152 * G + 0.0722 * B
def ratio(fg, bg):
    L1, L2 = lum(fg), lum(bg); hi, lo = max(L1, L2), min(L1, L2); return (hi + 0.05) / (lo + 0.05)
pairs = [
    ("ink/white", "000000", "FFFFFF"), ("ink/surface", "000000", "F4F4F4"),
    ("mute/white", "3B3B3B", "FFFFFF"), ("mute/surface", "3B3B3B", "F4F4F4"),
    ("mute/surface_deep", "3B3B3B", "EAEAEA"),
    ("ink/done", "000000", "ECECEC"), ("ink/wait", "000000", "EFE2E5"),
    ("vinho_esc/alert", "3A1018", "E7D2D7"), ("white/vinho", "FFFFFF", "5A1E2A"),
    ("vinho/white", "5A1E2A", "FFFFFF"),
]
for nm, fg, bg in pairs:
    r = ratio(fg, bg); lvl = "AAA" if r >= 7 else ("AA" if r >= 4.5 else ("AA-large" if r >= 3 else "FAIL"))
    print(("  OK  " if r >= 4.5 else (" WARN " if r >= 3 else " FAIL ")) + "%-20s %.2f:1 (%s)" % (nm, r, lvl))
    if r < 3:
        fails.append("contraste %s %.2f" % (nm, r))
    elif r < 4.5:
        warns.append("contraste %s %.2f (AA-large só)" % (nm, r))

# ==========================================================================
# PARTE D — integridade do documento de especificação (psi_planner_spec.html)
# Confere que o spec gerado segue a estrutura do PDF de referência e que os
# recursos embutidos (mockups base64) estão íntegros e completos.
# ==========================================================================
print("\n===== PARTE D: INTEGRIDADE DO SPEC (psi_planner_spec.html) =====")
SPEC_HTML = os.path.join(DELIV, "psi_planner2_spec.html")
if not os.path.exists(SPEC_HTML):
    check(False, "psi_planner_spec.html existe")
else:
    html = open(SPEC_HTML, encoding="utf-8").read()
    # 8 mockups embutidos (1 por aba do spec), todos base64 PNG não-vazios
    import re as _re
    imgs = _re.findall(r'data:image/png;base64,([A-Za-z0-9+/=]+)', html)
    check(len(imgs) >= 8, "8 mockups embutidos (achados: %d)" % len(imgs))
    check(all(len(b) > 1000 for b in imgs), "todos os mockups base64 não-triviais")
    check(len(set(imgs)) == len(imgs), "mockups distintos (sem duplicata/imagem quebrada)")
    # marcadores estruturais do PDF de referência
    markers = ["Especificação Visual da Planilha", "Psi <span class=\"wine\">Planner 2</span>",
               "Estrutura geral", "Identidade visual", "Paleta oficial", "Tipografia",
               "Notas técnicas", "Saluti · Projeto Psi Planner 2"]
    for mk in markers:
        check(mk in html, "spec contém marcador: %r" % mk[:32])
    # 7 abas + dashboard financeiro = 8 seções de aba
    for aba in ["Menu Inicial", "Financeiro", "Pacientes", "Calendário",
                "Aniversariantes", "Relatórios", "Configurações"]:
        check(html.count(aba) >= 1, "spec documenta aba: %s" % aba)
    # 6 cores da paleta oficial
    for hexc in ["#000000", "#3B3B3B", "#A6A6A6", "#F4F4F4", "#5A1E2A", "#3A1018"]:
        check(hexc in html, "paleta contém %s" % hexc)

print("\n==== RESULTADO: %s | %d falha(s), %d aviso(s) ====" %
      ("PASS" if not fails else "FALHAS", len(fails), len(warns)))
for x in fails:
    print(" FAIL -", x)
for x in warns:
    print(" WARN -", x)
sys.exit(1 if fails else 0)
